import openpyxl as xl
from .Transactions import *

class Workbook:
    def __init__(self, filename='./Data/Finances.xlsx'):
        self.filename = filename
        self.wb = xl.load_workbook(filename)
        self.Summary = self.wb.get_sheet_by_name('Summary')
        self.Incomes = self.wb.get_sheet_by_name('Incomes')
        self.Expenses = self.wb.get_sheet_by_name('Expenses')
        self.History = self.wb.get_sheet_by_name('Time History')
        self.Gasoline = self.wb.get_sheet_by_name('Gasoline')

    def add_transaction(self, transaction):
        if isinstance(transaction, Expense):
            ws = self.Expenses
        elif isinstance(transaction, Income):
            ws = self.Incomes

        # Row to be filled
        n = len(ws['B']) + 1

        # Date
        ws[f'A{n}'].value = transaction.date
        # Value
        ws[f'B{n}'].value = transaction.value
        ws[f'B{n}'].style = 'Currency'
        # Category
        ws[f'C{n}'].value = transaction.category
        # Description
        ws[f'D{n}'].value = transaction.description
        # account
        ws[f'E{n}'].value = transaction.account
        # Observations
        ws[f'F{n}'].value = transaction.observations

        # Update Summary
        val = self.update_summary(transaction.account, transaction.signal * transaction.value)
        # Update History
        self.update_history(transaction.account, transaction.date, val)


    def update_summary(self, account, delta):
        # Find account in summary
        for row in self.Summary['A']:
            if row.value == account:
                r = row.row
                break
        else:
            raise(Exception('account not found'))

        # Update account value
        self.Summary[f'B{r}'].value += delta
        
        return self.Summary[f'B{r}'].value
        
    def update_history(self, account, date, value):
        n = len(self.History['C']) + 1

        self.History[f'A{n}'] = date
        self.History[f'B{n}'] = account
        self.History[f'C{n}'] = value
        self.History[f'C{n}'].style = 'Currency'

    def add_gasoline(self, date, kilometers, liters):
        ws = self.Gasoline
    	# Row to be filled
        n = len(ws['B']) + 1

        autonomy = kilometers/liters

        # Date
        ws[f'A{n}'].value = date
        # Kilometers
        ws[f'B{n}'].value = kilometers
        # Liters
        ws[f'C{n}'].value = liters
        # Autonomy
        ws[f'D{n}'].value = autonomy


    def save_and_quit(self):
        self.wb.save(self.filename)
        self.wb.close()

